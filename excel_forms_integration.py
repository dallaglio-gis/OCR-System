"""
Excel Forms Tab Integration Module
- Discovers rig sections on the 'Forms' sheet (classic path)
- Adds NEW: insert_drilling_data_new_sheet() to create a per-form sheet replica for testing
- Provides test_integration() for UI diagnostics
- Uses keep_vba + atomic save to avoid corrupting .xlsm files
"""

from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Optional, Any, List
import logging
import os
import tempfile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import RIG_MAPPINGS


# -------------------- helpers --------------------

def _canon(s: str) -> str:
    return (s or "").upper().strip().replace(" ", "").replace("_", "").replace("-", "")

def _load_xlsm(path: str):
    # preserve macros / vba
    return load_workbook(filename=path, keep_vba=True, read_only=False, data_only=False)

def _safe_save_xlsm(wb, dest_path: str):
    base_dir = os.path.dirname(dest_path) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsm", dir=base_dir)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, dest_path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


# -------------------- results --------------------

@dataclass
class ExcelInsertionResult:
    success: bool
    worksheet_name: str
    row_inserted: int
    message: str
    data_inserted: Dict[str, Any]
    rig_section: str = ""


# -------------------- main class --------------------

class FormsTabIntegration:
    """Integration module for Excel 'Forms' tab with rig-specific sections, plus a new-sheet testing path."""

    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        self.rig_sections: Dict[str, Dict[str, Any]] = {}
        # Discovery is optional for the new-sheet path, but we keep it for the classic path.
        try:
            self._discover_rig_sections()
        except Exception as e:
            self.logger.warning(f"Rig discovery skipped due to error: {e}")

    # ---------- canon/match ----------
    def _canon(self, s: str) -> str: 
        return _canon(s)

    def _canon_rig(self, rig_id: str) -> str:
        x = _canon(rig_id)
        for standard, variants in RIG_MAPPINGS.items():
            targets = {_canon(standard)} | {_canon(v) for v in variants}
            if x in targets or any(t in x or x in t for t in targets):
                return standard
        return str(rig_id).strip()

    def _is_rig_identifier(self, value: str) -> bool:
        if not value:
            return False
        v = _canon(value)
        for standard, variants in RIG_MAPPINGS.items():
            targets = {_canon(standard)} | {_canon(vv) for vv in variants}
            if any(t in v or v in t for t in targets):
                return True
        return False

    # ---------- discovery of classic 'Forms' sections ----------
    def _discover_rig_sections(self):
        wb = _load_xlsm(self.excel_file_path)
        if 'Forms' not in wb.sheetnames:
            wb.close()
            return
        ws = wb['Forms']
        # wide & deep scan to catch sections at CE/CF and beyond
        for r in range(1, min(1000, ws.max_row + 1)):
            for c in range(1, min(200, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if not val:
                    continue
                if self._is_rig_identifier(str(val)):
                    info = self._find_rig_data_area(ws, r, c)
                    if info:
                        canon_name = self._canon_rig(str(val))
                        self.rig_sections[canon_name] = info
        wb.close()

    def _is_data_header(self, value: str) -> bool:
        headers = [
            'DATE', 'BHID', 'HOLE', 'DEPTH FROM', 'DEPTH TO', 'FROM', 'TO',
            'DRILLED', 'METER', 'RECOVERY', 'RQD', 'GEOLOGY', 'ROCK',
            'COMMENT', 'NOTE', 'SHIFT', 'LEVEL'
        ]
        v = (value or '').upper().strip()
        return any(h in v for h in headers)

    def _find_rig_data_area(self, ws: Worksheet, rig_row: int, rig_col: int) -> Optional[Dict]:
        header_row = None
        data_start_row = None
        headers_found = []

        for rr in range(rig_row + 1, min(rig_row + 60, ws.max_row + 1)):
            row_hits = []
            left = max(1, rig_col - 30)
            right = min(ws.max_column, rig_col + 60)
            for cc in range(left, right + 1):
                val = ws.cell(row=rr, column=cc).value
                if val and self._is_data_header(str(val)):
                    row_hits.append({'header': str(val).strip(), 'col': cc})
            if len(row_hits) >= 2:  # relaxed threshold
                header_row = rr
                data_start_row = rr + 1
                headers_found = row_hits
                break

        if not header_row:
            return None

        colmap: Dict[str, int] = {}
        for h in headers_found:
            H = h['header'].upper()
            c = h['col']
            if 'DATE' in H:
                colmap['date'] = c
            elif 'BHID' in H or 'HOLE' in H:
                colmap['hole_id'] = c
            elif 'DEPTH FROM' in H or H == 'FROM' or 'FROM(' in H:
                colmap['from_depth'] = c
            elif 'DEPTH TO' in H or H == 'TO' or 'TO(' in H:
                colmap['to_depth'] = c
            elif 'DRILLED' in H or 'METER' in H:
                colmap['drilled_meters'] = c
            elif 'RECOVERY' in H:
                colmap['core_recovery'] = c
            elif 'RQD' in H:
                colmap['rqd'] = c
            elif 'GEOLOGY' in H or 'ROCK' in H:
                colmap['geology'] = c
            elif 'COMMENT' in H or 'NOTE' in H:
                colmap['comments'] = c

        return {
            'rig_row': rig_row,
            'rig_col': rig_col,
            'header_row': header_row,
            'data_start_row': data_start_row,
            'column_mapping': colmap,
            'headers_found': headers_found,
        }

    def _find_next_available_row_in_section(self, ws: Worksheet, section: Dict) -> int:
        start = section['data_start_row']
        cmap = section['column_mapping']
        check_col = cmap.get('date') or cmap.get('hole_id') or list(cmap.values())[0]
        for r in range(start, start + 300):
            v = ws.cell(row=r, column=check_col).value
            if v is None or str(v).strip() == "":
                return r
        return start + 300

    def _insert_row(self, ws: Worksheet, row: int, section: Dict, data: Dict[str, Any]) -> Dict[str, Any]:
        cmap = section['column_mapping']
        inserted: Dict[str, Any] = {}

        # Map pipeline keys → excel logical fields
        fm = {
            'date': 'date',
            'hole_id': 'hole_id',
            'from_depth': 'from_depth',
            'to_depth': 'to_depth',
            'drilled_meters': 'total_depth',   # pipeline uses total_depth
            'core_recovery': 'core_recovery',
            'rqd': 'rqd',
            'geology': 'geology',
            'comments': 'comments',
        }

        for excel_field, data_field in fm.items():
            if excel_field in cmap and data_field in data and data[data_field] is not None:
                col = cmap[excel_field]
                val = data[data_field]
                if excel_field == 'date' and isinstance(val, str):
                    try:
                        val = datetime.strptime(val.strip(), '%Y-%m-%d')
                    except Exception:
                        pass
                ws.cell(row=row, column=col, value=val)
                inserted[excel_field] = val

        return inserted

    # ---------- public classic path ----------
    def insert_drilling_data(self, validated_data: Dict[str, Any]) -> ExcelInsertionResult:
        try:
            rig_id = validated_data.get('rig_id')
            if not rig_id:
                return ExcelInsertionResult(False, 'Forms', 0, 'No rig ID provided', {}, rig_section='')
            canon_rig = self._canon_rig(rig_id)
            if canon_rig not in self.rig_sections:
                debug = {'asked_for': rig_id, 'canonical': canon_rig, 'available_sections': list(self.rig_sections.keys())}
                return ExcelInsertionResult(False, 'Forms', 0, f'No matching rig section. {debug}', {}, rig_section=canon_rig)

            wb = _load_xlsm(self.excel_file_path)
            ws = wb['Forms']
            section = self.rig_sections[canon_rig]
            row = self._find_next_available_row_in_section(ws, section)
            inserted = self._insert_row(ws, row, section, validated_data)
            _safe_save_xlsm(wb, self.excel_file_path)
            wb.close()
            return ExcelInsertionResult(True, 'Forms', row, f'Inserted for {canon_rig}', inserted, rig_section=canon_rig)
        except Exception as e:
            return ExcelInsertionResult(False, 'Forms', 0, f'Error inserting data: {e}', {}, rig_section='')

    # ---------- NEW: per-form test sheet ----------
    def insert_drilling_data_new_sheet(self, validated_data: Dict[str, Any]) -> ExcelInsertionResult:
        """
        Create a new worksheet (FormEntry, FormEntry_1, ...) and write all extracted fields,
        including activity durations as decimal hours. This does NOT alter your Forms tab.
        """
        try:
            wb = _load_xlsm(self.excel_file_path)
            # unique sheet name
            base = "FormEntry"
            name = base
            idx = 0
            existing = set(wb.sheetnames)
            while name in existing:
                idx += 1
                name = f"{base}_{idx}"
            ws = wb.create_sheet(title=name)

            # --- header area (replica-style, compact) ---
            def W(r, c, v):
                ws.cell(row=r, column=c, value=v)

            # Metadata
            W(1,1,"Date");              W(1,2, _coerce_date(validated_data.get("date")))
            W(2,1,"Shift");             W(2,2, validated_data.get("shift"))
            W(3,1,"Level");             W(3,2, validated_data.get("level"))
            W(4,1,"Cubby ID");          W(4,2, validated_data.get("cubby_id"))
            # BHID (Hole ID)
            hole = validated_data.get("hole_id") or validated_data.get("bhid")
            W(5,1,"BHID");              W(5,2, hole)
            # Rig
            W(6,1,"Rig ID");            W(6,2, validated_data.get("rig_id"))
            # Personnel / admin
            W(7,1,"Driller");           W(7,2, validated_data.get("driller"))
            W(8,1,"Geo Tech");          W(8,2, validated_data.get("geotech"))
            W(9,1,"Daily Budget (m)");  W(9,2, _coerce_float(validated_data.get("daily_budget")))
            W(10,1,"Comments");         W(10,2, validated_data.get("comments"))

            # Depths / meters
            W(12,1,"Depth From (m)");   W(12,2, _coerce_float(validated_data.get("from_depth")))
            W(13,1,"Depth To (m)");     W(13,2, _coerce_float(validated_data.get("to_depth")))
            meters = validated_data.get("total_depth", validated_data.get("drilling_meters"))
            W(14,1,"Drilled Meters");   W(14,2, _coerce_float(meters))
            W(15,1,"Core Recovery %");  W(15,2, _coerce_float(validated_data.get("core_recovery")))
            W(16,1,"RQD %");            W(16,2, _coerce_float(validated_data.get("rqd")))
            W(17,1,"Geology");          W(17,2, validated_data.get("geology"))

            # Activity Record table (from, to, text, duration)
            W(19,1,"Activity Record")
            W(20,1,"Row"); W(20,2,"From"); W(20,3,"To"); W(20,4,"Activity"); W(20,5,"Duration (h)")

            rows = validated_data.get("activity_rows") or []
            r = 21; tot = 0.0
            for i, row_data in enumerate(rows, start=1):
                W(r,1,i)
                W(r,2,row_data.get("from"))
                W(r,3,row_data.get("to"))
                W(r,4,row_data.get("text"))
                dur = row_data.get("duration")
                W(r,5, dur)
                if isinstance(dur, (int,float)): tot += float(dur)
                r += 1
            W(r,4,"Total")
            W(r,5, round(tot,3))

            _safe_save_xlsm(wb, self.excel_file_path)
            wb.close()

            return ExcelInsertionResult(True, name, r, f"Created new sheet '{name}'", {
                "date": validated_data.get("date"),
                "rig_id": validated_data.get("rig_id"),
                "bhid": hole,
                "durations_count": len(rows),
                "durations_total": round(tot, 3),
            })
        except Exception as e:
            return ExcelInsertionResult(False, "", 0, f"Error creating test sheet: {e}", {})

    # ---------- diagnostics ----------
    def test_integration(self) -> Dict[str, Any]:
        """Summarise discovered rig sections for debugging in the UI."""
        try:
            return {
                "excel_file_path": self.excel_file_path,
                "rig_sections_found": len(self.rig_sections),
                "rig_sections": {
                    name: {
                        "rig_row": info.get("rig_row"),
                        "rig_col": info.get("rig_col"),
                        "header_row": info.get("header_row"),
                        "data_start_row": info.get("data_start_row"),
                        "columns_found": len(info.get("column_mapping", {})),
                        "column_mapping": info.get("column_mapping", {}),
                    }
                    for name, info in self.rig_sections.items()
                },
            }
        except Exception as e:
            return {"error": f"test_integration failed: {e}"}


# -------------------- small coercers --------------------

def _coerce_date(v: Any):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        return v  # leave as-is if unknown
    return v

def _coerce_float(v: Any):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return v
